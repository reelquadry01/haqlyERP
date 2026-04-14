# Author: Quadri Atharu
"""Excel export engine using openpyxl with financial formatting."""

from __future__ import annotations

import os
from datetime import datetime
from typing import Any, Dict, List, Optional, Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side, numbers
from openpyxl.utils import get_column_letter

from ..core.logging import get_logger

logger = get_logger(__name__)

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
TOTAL_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
TOTAL_FONT = Font(name="Calibri", bold=True, size=11)
NUMBER_FORMAT_NGN = '#,##0.00'
NUMBER_FORMAT_PCT = '0.00%'
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
BOTTOM_DOUBLE = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="double"),
)


def export_to_excel(
    data: List[Dict[str, Any]],
    filename: str,
    sheet_name: str = "Sheet1",
    output_dir: str = ".",
) -> str:
    if not data:
        logger.warning("export_to_excel_empty_data", filename=filename)
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        ws["A1"] = "No data to export"
        filepath = os.path.join(output_dir, _ensure_extension(filename, ".xlsx"))
        wb.save(filepath)
        return filepath

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]

    headers = list(data[0].keys())
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=_format_header(header))
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER

    for row_idx, record in enumerate(data, 2):
        for col_idx, header in enumerate(headers, 1):
            value = record.get(header)
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center")

            if isinstance(value, float) or isinstance(value, int):
                if isinstance(value, float) and abs(value) < 1 and value != 0:
                    cell.number_format = NUMBER_FORMAT_PCT
                else:
                    cell.number_format = NUMBER_FORMAT_NGN
            elif isinstance(value, datetime):
                cell.number_format = "YYYY-MM-DD HH:MM:SS"

    format_financial_sheet(ws, data)

    for col_idx, header in enumerate(headers, 1):
        max_length = len(str(_format_header(header)))
        for row in ws.iter_rows(min_row=2, max_row=min(len(data) + 1, 50), min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 4, 40)

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"

    filepath = os.path.join(output_dir, _ensure_extension(filename, ".xlsx"))
    wb.save(filepath)
    logger.info("excel_exported", filename=filepath, rows=len(data), columns=len(headers))
    return filepath


def format_financial_sheet(ws: Any, data: List[Dict[str, Any]]) -> None:
    if not data:
        return

    headers = list(data[0].keys())
    numeric_cols: List[int] = []
    for col_idx, header in enumerate(headers, 1):
        values = [r.get(header) for r in data if r.get(header) is not None]
        if values and all(isinstance(v, (int, float)) for v in values):
            numeric_cols.append(col_idx)

    if numeric_cols:
        total_row = len(data) + 2
        ws.cell(row=total_row, column=1, value="TOTAL").font = TOTAL_FONT
        ws.cell(row=total_row, column=1).fill = TOTAL_FILL
        ws.cell(row=total_row, column=1).border = BOTTOM_DOUBLE

        for col_idx in numeric_cols:
            col_letter = get_column_letter(col_idx)
            total_cell = ws.cell(row=total_row, column=col_idx)
            total_cell.value = f"=SUM({col_letter}2:{col_letter}{len(data) + 1})"
            total_cell.number_format = NUMBER_FORMAT_NGN
            total_cell.font = TOTAL_FONT
            total_cell.fill = TOTAL_FILL
            total_cell.border = BOTTOM_DOUBLE
            total_cell.alignment = Alignment(horizontal="right")

        for col_idx in range(2, len(headers) + 1):
            if col_idx not in numeric_cols:
                cell = ws.cell(row=total_row, column=col_idx)
                cell.fill = TOTAL_FILL
                cell.border = BOTTOM_DOUBLE

    for row in ws.iter_rows(min_row=2, max_row=len(data) + 1):
        for cell in row:
            if cell.number_format == "General" and isinstance(cell.value, (int, float)):
                cell.number_format = NUMBER_FORMAT_NGN


def _format_header(header: str) -> str:
    return header.replace("_", " ").title()


def _ensure_extension(filename: str, ext: str) -> str:
    if not filename.lower().endswith(ext):
        return filename + ext
    return filename
